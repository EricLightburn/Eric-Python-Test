import os
from openpyxl import Workbook, load_workbook
import sys

class Color:
    '''Defines colour codes for use in terminal'''
    RED = '\033[31m'
    GREEN = '\033[32m'
    YELLOW = '\033[33m'
    BLUE = '\033[34m'
    MAGENTA = '\033[35m'
    CYAN = '\033[36m'
    WHITE = '\033[37m'
    RESET = '\033[0m'
class Record:
    '''Represents a single record with name, age, and interests'''
    def __init__(self, name, age, interests):
        self.name = name
        self.age = age
        self.interests = interests
class RecordFile:
    '''Manages reading and writing records to the file'''
    def __init__(self, file_path):
        self.file_path = file_path
        if not os.path.exists(self.file_path):
            self.create_file()
    def create_file(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '#'
        ws['B1'] = 'Name'
        ws['C1'] = 'Age'
        ws['D1'] = 'Interests'
        wb.save(self.file_path)
    def read_records(self):
        '''Reads records from the file and returns them as a list of Record objects'''
        records = []
        wb = load_workbook(self.file_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            record = Record(row[1].value, row[2].value, row[3].value)
            records.append(record)
        return records
    def write_record(self, record_count, record):
        '''Writes a single record to the file'''
        wb = load_workbook(self.file_path)
        ws = wb.active
        ws.append([record_count, record.name, record.age, record.interests])
        wb.save(self.file_path)
class RecordManager:
    '''Provides an interface for adding, viewing, deleting, clearing and finding records'''
    def __init__(self):
        file_path = os.path.join("record.xlsx")
        self.record_file = RecordFile(file_path)
        self.records = self.record_file.read_records()
        self.record_count = len(self.records)
    def add_record(self):
        '''Prompts the user for info and adds it to a new record'''
        name = input("Enter your name: ")
        age = input("Enter your age: ")
        interests = input("Enter your interest: ")
        record = Record(name, age, interests)
        self.record_count += 1
        self.records.append(record)
        self.record_file.write_record(self.record_count, record)
    def view_records(self):
        '''Displays all records in terminal'''
        if len(self.records) == 0:
            print(f"{Color.YELLOW}There are no records{Color.RESET}")
            print()
        else:
            wb = load_workbook(self.record_file.file_path)
            ws = wb.active
            for row in ws.iter_rows():
                for cell in row:
                    print(cell.value, end='\t')
                print()
    def clear_records(self):
        '''Clears all records from the file and record list'''
        self.records = []
        with open(self.record_file.file_path, "w") as f:
            f.truncate(0)
            f.write(f"{Color.BLUE}#\t\t\tName\t\t\t\tAge\t\t\t\tInterests{Color.RESET}\n")
    def delete_record(self):
        '''Prompts the user for a record number and deletes the one chosen'''
        try:
            record_id = int(input("Enter Record #: ")) - 1
            if not (0 <= record_id < len(self.records)):
                raise ValueError("Invalid record number.")
            del self.records[record_id]
            self.record_count -= 1
            self.record_file.create_file()
            for i, record in enumerate(self.records):
                self.record_file.write_record(i+1, record)
            print(f"{Color.GREEN}Record deleted successfully.{Color.RESET}")
        except ValueError as e:
            print(f"{Color.RED}{e}{Color.RESET}")
    def find_records(self):
        '''Prompts the user to search for info'''
        name = input("Enter name to search for: ")
        wb = load_workbook(self.record_file.file_path)
        ws = wb.active
        for row in ws.iter_rows():
            if row[1].value == name:
                for cell in row:
                    print(cell.value, end='\t')
                    print()
manager = RecordManager()
while True:            
    print(f"{Color.CYAN}=========================================================={Color.RESET}")
    print(f"{Color.CYAN} ~ ----------------- Record Manager --------------------- ~{Color.RESET}")
    print(f"{Color.CYAN}==========================================================={Color.RESET}")
    print("Available Options:")
    print(f"A) {Color.GREEN}Add Record{Color.RESET}\tB) {Color.YELLOW}View Record{Color.RESET}")
    print(f"C) {Color.RED}Delete Record{Color.RESET}\tD) {Color.MAGENTA}Clear Records{Color.RESET}")
    print(f"E) {Color.WHITE}Exit Manager{Color.RESET}\tF) {Color.BLUE}Find Records{Color.RESET} \n")
    selection = input('Enter selection: ').upper()
    if selection == 'A':
        manager.add_record()
    elif selection == 'B':
        manager.view_records()
    elif selection == 'C':
        manager.delete_record()
    elif selection == 'D':
        manager.clear_records()
    elif selection == 'F':
        manager.find_records()
    elif selection == 'E':
        sys.exit()
    else:
        print("Invalid selection. Please try again.")
