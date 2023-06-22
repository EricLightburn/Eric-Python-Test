import os
import sys
from openpyxl import load_workbook

file_path = r"c:\Users\training.user\Downloads\Book1.xlsx"

def add_record():
    global file_path
    if os.path.exists(file_path):
        workbook = load_workbook(filename=file_path)
        worksheet = workbook.active
    else:
        print("File not found.")

    name = input("Enter your name: ")
    age = input("Enter your age: ")
    interests = input("Enter your interest: ")

    row = worksheet.max_row + 1
    worksheet.cell(row=row, column=1, value=name)
    worksheet.cell(row=row, column=2, value=age)
    worksheet.cell(row=row, column=3, value=interests)

    workbook.save(filename=file_path)

def view_records():
    global file_path
    workbook = load_workbook(filename=file_path)
    worksheet = workbook.active

    for row in worksheet.iter_rows(values_only=True):
        print(row)

def delete_record(line_number):
    global file_path
    workbook = load_workbook(filename=file_path)
    worksheet = workbook.active

    worksheet.delete_rows(line_number)

    workbook.save(filename=file_path)

    print("Record Deleted.")

def clear_records():
    global file_path
    workbook = load_workbook(filename=file_path)
    worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=2):
        worksheet.delete_rows(row[0].row)

    workbook.save(filename=file_path)

while True:
    print("==========================================================")
    print(" ~ ----------------- Record Manager --------------------- ~")
    print("===========================================================")
    print("Available Options:")
    print("A) Add Record\t\tB) View Record")
    print("C) Clear Records\tD) Delete Record")
    print("E. Exit\n")

    choice = input('Enter selection: ').upper()

    if choice == "A":
        add_record()
    elif choice == "B":
        view_records()
    elif choice == "C":
        clear_records()
    elif choice == "D":
        delete_record(int(input("Which row to delete? ")))
    elif choice == "E":
        sys.exit("Thank you!")
    else:
        print("Invalid response. Please try again.")
